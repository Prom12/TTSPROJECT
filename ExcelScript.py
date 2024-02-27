import openpyxl

from SharedGlobalVariables import headers,Singleton


def open_and_load_workbook(filepath):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    return wb,sheet

def display_data(sheet):
    data = []
    for row in sheet.iter_rows(values_only = True):
        data.append(row)
    return data

def read_excel_and_display(filepath):
    # Open excel
    wb,sheet = open_and_load_workbook(filepath)

    # Read data from the Excel sheet
    return display_data(sheet)

def read_recorder_and_display(filepath):
    # Open excel
    wb,sheet = open_and_load_workbook(filepath)
    # create sheet
    if Singleton.username != '':
       wb,sheet = open_or_create_sheet(wb,filepath)
       Singleton.activeWB = wb
       Singleton.activeSheet = sheet
    else:
        print('No User Exists')
    
    return display_data(sheet)
        
def open_or_create_sheet(wb,filepath):
    if Singleton.username in wb.sheetnames:
        sheet = wb[Singleton.username]
    else:
        sheet = wb.create_sheet(Singleton.username)

        # Add Headers
        for col, header in enumerate(headers,start=1):
            sheet.cell(row=1, column =col, value = header)
    wb.save(filepath)
    
    return wb,sheet
